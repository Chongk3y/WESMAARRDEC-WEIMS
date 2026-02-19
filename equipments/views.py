# Standard library imports
import csv
import json
import traceback
from datetime import datetime
from functools import wraps
# Django core imports
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, Http404
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Count, F, Q, Sum
from django.core.exceptions import ObjectDoesNotExist
# Django auth imports
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from auth_user.models import User
# Third-party imports
import openpyxl
from django.db.models.functions import TruncMonth, ExtractYear

# Import WESMAARRDEC auth system
from cmscore.decorators import secretariat_required

# Local app imports
from .models import Equipment, Category, Status
from .models import EquipmentHistory, ReplacementDocument
from .models import EquipmentActionLog
from .models import ReportTemplate
from .models import ReturnDocument, ReplacementDocument
from .forms import ReportFilterForm
from .helpers import (
    is_admin,
    is_encoder,
    is_client,
    is_superadmin,
    is_admin_or_superadmin,
    is_admin_superadmin_encoder,
    encoder_readonly_or_denied,
    is_viewer_or_above,
    role_required_with_feedback
)

@login_required
def equipment_table_json(request):
    try:
        is_client_user = is_client(request.user)

        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        # Initial queryset
        qs = Equipment.objects.filter(is_returned=False, is_archived=False).select_related('category', 'status', 'emp')

        # Global search
        if search_value:
            qs = qs.filter(
                Q(item_propertynum__icontains=search_value) |
                Q(item_name__icontains=search_value) |
                Q(item_desc__icontains=search_value) |
                Q(category__name__icontains=search_value) |
                Q(status__name__icontains=search_value) |
                Q(po_number__icontains=search_value) |
                Q(fund_source__icontains=search_value) |
                Q(supplier__icontains=search_value) |
                Q(item_amount__icontains=search_value) |
                Q(assigned_to__icontains=search_value) |
                Q(end_user__icontains=search_value) |
                Q(location__icontains=search_value) |
                Q(current_location__icontains=search_value) |
                Q(item_purdate__icontains=search_value) |
                Q(project_name__icontains=search_value)
            )

        # Advanced filters
        for key, value in request.GET.items():
            if key.startswith('filter_col_') and value:
                col_idx = key.replace('filter_col_', '')
                if col_idx == '2':  # Property #
                    qs = qs.filter(item_propertynum__icontains=value)
                elif col_idx == '3':  # Name
                    qs = qs.filter(item_name__icontains=value)
                elif col_idx == '4':  # Description
                    qs = qs.filter(item_desc__icontains=value)
                elif col_idx == '5':  # PO Number
                    qs = qs.filter(po_number__icontains=value)
                elif col_idx == '6':  # Fund Source
                    qs = qs.filter(fund_source__icontains=value)
                elif col_idx == '7':  # Supplier
                    qs = qs.filter(supplier__icontains=value)
                elif col_idx == '8':  # Amount
                    try:
                        qs = qs.filter(item_amount=float(value))
                    except Exception:
                        pass
                elif col_idx == '9':  # Assigned To
                    qs = qs.filter(assigned_to__icontains=value)
                elif col_idx == '10':  # End User
                    qs = qs.filter(end_user__icontains=value)
                elif col_idx == '11':  # Location
                    qs = qs.filter(location__icontains=value)
                elif col_idx == '12':  # Current Location
                    qs = qs.filter(current_location__icontains=value)
                elif col_idx == '13':  # Category
                    qs = qs.filter(category__name__icontains=value)
                elif col_idx == '14':  # Status
                    qs = qs.filter(status__name__icontains=value)
                elif col_idx == '15':  # Purchase Date
                    qs = qs.filter(item_purdate=value)

        # Sorting
        order_col = request.GET.get('order[0][column]', '1')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        col_map = {
            '1': 'id',
            '3': 'item_propertynum',
            '4': 'item_name',
            '5': 'item_desc',
            '6': 'po_number',
            '7': 'item_amount',
            '8': 'end_user',
            '9': 'category__name',
            '10': 'status__name',
        }

        order_field = col_map.get(order_col, 'id')
        if order_dir == 'desc':
            order_field = '-' + order_field

        total = Equipment.objects.filter(is_returned=False).count()
        filtered = qs.count()

        equipments = qs.order_by(order_field)[start:start + length]

        data = []
        for eq in equipments:
            
            # Only show Delete and Archive for admin/superadmin
            actions = ''
            if not is_client(request.user):  # ✅ hide entire dropdown for clients
                actions = f'''
                <div class="dropdown" data-bs-auto-close="outside">
                <button class="btn btn-outline-secondary btn-sm dropdown-toggle" type="button" data-bs-toggle="dropdown" aria-expanded="false">
                    Actions
                </button>
                <ul class="dropdown-menu">
                    <li>
                    <a class="dropdown-item" href="/weims/edit/{eq.id}/">
                        <i class="bi bi-pencil-square"></i> Edit
                    </a>
                    </li>
                '''
                if is_admin(request.user) or is_superadmin(request.user):
                    safe_name = (eq.item_name or 'Unknown Equipment').replace("'", "\\'")
                    safe_property = eq.item_propertynum or 'N/A'
                    actions += f'''
                    <li>
                    <a class="dropdown-item" href="#" onclick="showIndividualDeleteConfirm({eq.id}, '{safe_name} - #{safe_property}', '/weims/delete/{eq.id}/');">
                        <i class="bi bi-trash"></i> Delete
                    </a>
                    </li>
                    '''
                if is_admin(request.user) or is_superadmin(request.user) or is_encoder(request.user):
                    safe_name = (eq.item_name or 'Unknown Equipment').replace("'", "\\'")
                    safe_property = eq.item_propertynum or 'N/A'
                    actions += f'''
                    <li>
                    <a class="dropdown-item" href="#" onclick="showIndividualArchiveConfirm({eq.id}, '{safe_name} - #{safe_property}', '/weims/archive/{eq.id}/');">
                        <i class="bi bi-archive"></i> Archive
                    </a>
                    </li>
                    '''
                if not eq.is_returned:
                    if eq.status and eq.status.name == 'Lost':
                        # Show disabled button with tooltip for lost equipment
                        actions += f'''
                        <li>
                        <button type="button" class="dropdown-item disabled" disabled 
                                data-bs-toggle="tooltip" data-bs-placement="top" 
                                title="Lost equipment cannot be returned">
                            <i class="bi bi-arrow-90deg-left text-muted"></i> Return
                        </button>
                        </li>
                        '''
                    else:
                        # Show normal return button
                        actions += f'''
                        <li>
                        <button type="button" class="dropdown-item" data-bs-toggle="modal" data-bs-target="#returnModal" 
                                data-eqid="{eq.id}" 
                                data-eqname="{eq.item_name or 'N/A'}" 
                                data-eqproperty="{eq.item_propertynum or 'N/A'}">
                            <i class="bi bi-arrow-90deg-left"></i> Return
                        </button>
                        </li>
                        '''
                actions += '''
                </ul>
                </div>
                '''
            status_colors = {
                'Active': 'success',
                'Damaged': 'danger',
                'Maintenance': 'warning',
                'Lost': 'secondary',
                'Condemned': 'dark',
            }
            row = [
                '',  # 0: Checkbox placeholder
                eq.id,  # 1: ID
                f'<img src="{eq.user_image.url if eq.user_image else ""}" style="width:32px;height:32px;object-fit:cover;" class="img-thumbnail">',  # 2: Image
                eq.item_propertynum or '',  # 3: Property #
                eq.item_name or '',  # 4: Name
                eq.item_desc if eq.item_desc else 'None',  # 5: Description
                eq.po_number if eq.po_number else 'None',  # 6: PO Number
                f'₱{eq.item_amount:,.2f}' if eq.item_amount else '₱0.00',  # 7: Amount
                eq.end_user if eq.end_user else 'None',  # 8: End User
                eq.category.name if eq.category else 'None',  # 9: Category
                f'<span class="badge bg-{status_colors.get(eq.status.name if eq.status else "secondary", "secondary")}">{eq.status.name if eq.status else "None"}</span>',  # 10: Status
                actions if not is_client_user else '',  # 11: Actions - always include
                # Hidden columns
                eq.fund_source if eq.fund_source else 'None',  # 12: Fund Source
                eq.supplier if eq.supplier else 'None',  # 13: Supplier
                eq.assigned_to if eq.assigned_to else 'None',  # 14: Assigned To
                eq.location if eq.location else 'None',  # 15: Deployment Location
                eq.current_location if eq.current_location else 'None',  # 16: Current Location
                eq.item_purdate.strftime('%Y-%m-%d') if eq.item_purdate else 'None',  # 17: PO Date
                eq.project_name if eq.project_name else 'None',  # 18: Project Name
            ]

            data.append(row)

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total,
            'recordsFiltered': filtered,
            'data': data,
        })
        
    except Exception as e:
        import traceback
        # Return error details for debugging
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'draw': request.GET.get('draw', 1),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        }, status=500)


@login_required
def equipment_detail_json(request, pk):
    try:
        eq = Equipment.objects.select_related('category', 'status', 'created_by', 'updated_by').get(pk=pk)
    except Equipment.DoesNotExist:
        raise Http404
    
    # Get replacement documents for lost equipment
    replacement_docs = []
    if eq.status and eq.status.name == 'Lost':
        replacement_documents = eq.replacement_documents.all().order_by('-uploaded_at')
        for doc in replacement_documents:
            replacement_docs.append({
                'url': doc.document.url,
                'filename': doc.original_filename or doc.document.name.split('/')[-1],
                'uploaded_at': doc.uploaded_at.strftime("%b %d, %Y %H:%M"),
                'uploaded_by': doc.uploaded_by.get_full_name() if doc.uploaded_by else 'Unknown'
            })
    
    data = {
        "image": eq.user_image.url if eq.user_image else "",
        "propertynum": eq.item_propertynum or "None",
        "name": eq.item_name or "None",
        "desc": eq.item_desc if eq.item_desc not in (None, '') else "None",
        "addinfo": eq.additional_info if eq.additional_info not in (None, '') else "None",
        "amount": f"{eq.item_amount:,.2f}" if eq.item_amount is not None else "None",
        "category": eq.category.name if eq.category and eq.category.name else "None",
        "status": eq.status.name if eq.status and eq.status.name else "None",
        "damage_reason": eq.damage_reason if eq.damage_reason not in (None, '') else "None",
        "lost_remarks": eq.lost_remarks if eq.lost_remarks not in (None, '') else "None",
        "replacement_documents": replacement_docs,
        "po_number": eq.po_number if eq.po_number not in (None, '') else "None",
        "fund_source": eq.fund_source if eq.fund_source not in (None, '') else "None",
        "supplier": eq.supplier if eq.supplier not in (None, '') else "None",
        "po_date": eq.item_purdate.strftime("%b %d, %Y") if eq.item_purdate else "None",
        "project_name": eq.project_name if eq.project_name not in (None, '') else "None",
        "assigned_to": eq.assigned_to if eq.assigned_to not in (None, '') else "None",
        "end_user": eq.end_user if eq.end_user not in (None, '') else "None",
        "location": eq.location if eq.location not in (None, '') else "None",
        "current_location": eq.current_location if eq.current_location not in (None, '') else "None",
        "created": eq.created_at.strftime("%b %d, %Y") if eq.created_at else "None",
        "created_by": f"{eq.created_by.first_name} {eq.created_by.last_name}" if eq.created_by else "None",
        "updated": eq.updated_at.strftime("%b %d, %Y") if eq.updated_at else "None",
        "updated_by": f"{eq.updated_by.first_name} {eq.updated_by.last_name}" if eq.updated_by else "None",
    }
    return JsonResponse(data)

@login_required
@role_required_with_feedback(is_viewer_or_above)
@secretariat_required
def index(request):
    
    equipments = Equipment.objects.filter(is_returned=False).select_related('category', 'status', 'emp').all()
    categories = Category.objects.all()
    statuses = Status.objects.all()
    end_users = Equipment.objects.exclude(end_user__isnull=True).exclude(end_user='').values_list('end_user', flat=True).distinct()
    assigned_to_list = Equipment.objects.exclude(assigned_to__isnull=True).exclude(assigned_to='').values_list('assigned_to', flat=True).distinct()
    fund_sources = Equipment.objects.exclude(fund_source__isnull=True).exclude(fund_source='').values_list('fund_source', flat=True).distinct()
    suppliers = Equipment.objects.exclude(supplier__isnull=True).exclude(supplier='').values_list('supplier', flat=True).distinct()
    locations = Equipment.objects.exclude(location__isnull=True).exclude(location='').values_list('location', flat=True).distinct()
    category_id = request.GET.get('category')
    status_id = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    order = request.GET.get('order', 'desc')
    equipments = equipments.order_by('-item_purdate' if order == 'desc' else 'item_purdate')

    if category_id:
        equipments = equipments.filter(category_id=category_id)
    if status_id:
        equipments = equipments.filter(status_id=status_id)
    if date_from:
        equipments = equipments.filter(item_purdate__gte=date_from)
    if date_to:
        equipments = equipments.filter(item_purdate__lte=date_to)

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(equipments, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'equipments': page_obj,
        'categories': categories,
        'statuses': statuses,
        'end_users': end_users,
        'assigned_to_list': assigned_to_list,
        'fund_sources': fund_sources,
        'suppliers': suppliers,
        'locations': locations,
        'selected_category': category_id,
        'selected_status': status_id,
        'date_from': date_from,
        'date_to': date_to,
        'is_admin': is_admin(request.user),
        'is_encoder': is_encoder(request.user),
        'is_client': is_client(request.user),
        'import_error': request.session.pop('import_error', None),
    }
    
    return render(request, 'equipments/equipment_list.html', context)


@login_required
@user_passes_test(is_admin_superadmin_encoder)
def add_equipment(request):
    users = User.objects.all()
    categories = Category.objects.all()
    statuses = Status.objects.all()
    # Get distinct values from existing equipment for combo boxes
    existing_end_users = Equipment.objects.exclude(end_user__isnull=True).exclude(end_user__exact='').values_list('end_user', flat=True).distinct().order_by('end_user')
    existing_assigned_to = Equipment.objects.exclude(assigned_to__isnull=True).exclude(assigned_to__exact='').values_list('assigned_to', flat=True).distinct().order_by('assigned_to')
    existing_project_names = Equipment.objects.exclude(project_name__isnull=True).exclude(project_name__exact='').values_list('project_name', flat=True).distinct().order_by('project_name')
    existing_locations = Equipment.objects.exclude(location__isnull=True).exclude(location__exact='').values_list('location', flat=True).distinct().order_by('location')
    existing_current_locations = Equipment.objects.exclude(current_location__isnull=True).exclude(current_location__exact='').values_list('current_location', flat=True).distinct().order_by('current_location')
    from datetime import date
    today_date = date.today()
    return render(request, 'equipments/add.html', {
        'users': users,
        'categories': categories,
        'statuses': statuses,
        'existing_end_users': existing_end_users,
        'existing_assigned_to': existing_assigned_to,
        'existing_project_names': existing_project_names,
        'existing_locations': existing_locations,
        'existing_current_locations': existing_current_locations,
        'today_date': today_date,
        'is_admin': is_admin(request.user),
        'is_encoder': is_encoder(request.user),
    })

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def processaddequipment(request):
    errors = {}
    values = request.POST
    if request.method == 'POST':
        item_propertynum = request.POST.get('item_propertynum')
        item_name = request.POST.get('item_name')
        item_desc = request.POST.get('item_desc') or None
        item_purdate = request.POST.get('item_purdate') or None
        po_number = request.POST.get('po_number')
        fund_source = request.POST.get('fund_source') or None
        supplier = request.POST.get('supplier') or None
        item_amount = request.POST.get('item_amount')
        assigned_to = request.POST.get('assigned_to') or None
        location = request.POST.get('location') or None
        current_location = request.POST.get('current_location') or None
        project_name = request.POST.get('project_name') or None
        end_user = request.POST.get('end_user') or None
        category_id = request.POST.get('category_id')
        status_id = request.POST.get('status_id')
        damage_reason = request.POST.get('damage_reason') or None
        lost_remarks = request.POST.get('lost_remarks') or None
        user_image = request.FILES.get('user_image', 'equipment_pic/image.jpg')
        order_receipt = request.FILES.get('order_receipt', None)
        replacement_documents = request.FILES.getlist('replacement_documents')
        
        # Handle reason fields (usually empty for new equipment)
        assignment_reason = request.POST.get('assignment_reason') or None
        user_change_reason = request.POST.get('user_change_reason') or None
        location_change_reason = request.POST.get('location_change_reason') or None
        current_location_reason = request.POST.get('current_location_reason') or None

        # Field validations (keep as is)
        # ...existing validation code...

        # Validate item_amount is a valid number
        try:
            if item_amount is not None and item_amount != '':
                item_amount = float(item_amount)
        except ValueError:
            errors['item_amount'] = 'Amount must be a valid number.'

        if errors:
            users = User.objects.all()
            categories = Category.objects.all()
            statuses = Status.objects.all()
            # Re-fetch existing values for combo boxes when there are errors
            existing_end_users = Equipment.objects.exclude(end_user__isnull=True).exclude(end_user__exact='').values_list('end_user', flat=True).distinct().order_by('end_user')
            existing_assigned_to = Equipment.objects.exclude(assigned_to__isnull=True).exclude(assigned_to__exact='').values_list('assigned_to', flat=True).distinct().order_by('assigned_to')
            existing_project_names = Equipment.objects.exclude(project_name__isnull=True).exclude(project_name__exact='').values_list('project_name', flat=True).distinct().order_by('project_name')
            existing_locations = Equipment.objects.exclude(location__isnull=True).exclude(location__exact='').values_list('location', flat=True).distinct().order_by('location')
            existing_current_locations = Equipment.objects.exclude(current_location__isnull=True).exclude(current_location__exact='').values_list('current_location', flat=True).distinct().order_by('current_location')
            from datetime import date
            today_date = date.today()
            return render(request, 'equipments/add.html', {
                'errors': errors,
                'values': values,
                'users': users,
                'categories': categories,
                'statuses': statuses,
                'existing_end_users': existing_end_users,
                'existing_assigned_to': existing_assigned_to,
                'existing_project_names': existing_project_names,
                'existing_locations': existing_locations,
                'existing_current_locations': existing_current_locations,
                'today_date': today_date,
                'is_admin': is_admin(request.user),
                'is_encoder': is_encoder(request.user),
            })
        else:
            equipment = Equipment.objects.create(
                user_image=user_image,
                item_propertynum=item_propertynum,
                item_name=item_name,
                item_desc=item_desc,
                item_purdate=item_purdate if item_purdate else None,
                po_number=po_number,
                fund_source=fund_source,
                supplier=supplier,
                item_amount=item_amount,
                assigned_to=assigned_to,
                location=location,
                current_location=current_location,
                project_name=project_name,
                end_user=end_user,
                emp=request.user,
                category_id=category_id,
                status_id=status_id,
                damage_reason=damage_reason,
                lost_remarks=lost_remarks,
                assignment_reason=assignment_reason,
                user_change_reason=user_change_reason,
                location_change_reason=location_change_reason,
                current_location_reason=current_location_reason,
                created_by=request.user,
                updated_by=request.user,
                order_receipt=order_receipt 
            )
            equipment.save()
            
            # Handle replacement documents if any
            for doc in replacement_documents:
                ReplacementDocument.objects.create(
                    equipment=equipment,
                    document=doc,
                    original_filename=doc.name,
                    uploaded_by=request.user
                )
            EquipmentActionLog.objects.create(
            equipment=equipment,
            action='create',
            user=request.user,
            summary=f"Created equipment: {equipment.item_name} (Property #: {equipment.item_propertynum})"
)
            return HttpResponseRedirect('/weims/')
      
@login_required 
@user_passes_test(is_admin_superadmin_encoder)
def edit_equipment(request, id):
    equipment = get_object_or_404(Equipment, id=id)
    categories = Category.objects.all()
    statuses = Status.objects.all()
    users = User.objects.all()

    if request.method == 'POST':
        # Capture current values before changes (excluding reason fields to avoid tracking them)
        original = {
            'item_propertynum': equipment.item_propertynum,
            'item_name': equipment.item_name,
            'item_desc': equipment.item_desc,
            'item_purdate': equipment.item_purdate,
            'po_number': equipment.po_number,
            'fund_source': equipment.fund_source,
            'supplier': equipment.supplier,
            'item_amount': equipment.item_amount,
            'assigned_to': equipment.assigned_to,
            'location': equipment.location,
            'current_location': equipment.current_location,
            'project_name': equipment.project_name,
            'end_user': equipment.end_user,
            'emp_id': equipment.emp_id,
            'category_id': equipment.category_id,
            'status_id': equipment.status_id,
            'damage_reason': equipment.damage_reason,
            'lost_remarks': equipment.lost_remarks,
            # Reason fields excluded from history tracking
            # 'assignment_reason': equipment.assignment_reason,
            # 'user_change_reason': equipment.user_change_reason,
            # 'location_change_reason': equipment.location_change_reason,
            # 'current_location_reason': equipment.current_location_reason,
        }

        equipment.item_propertynum = request.POST.get('item_propertynum')
        equipment.item_name = request.POST.get('item_name')
        equipment.item_desc = request.POST.get('item_desc')

        # Handle item_purdate safely
        item_purdate = request.POST.get('item_purdate')
        if item_purdate:
            try:
                equipment.item_purdate = datetime.strptime(item_purdate, '%Y-%m-%d').date()
            except ValueError:
                equipment.item_purdate = None
        else:
            equipment.item_purdate = None

        equipment.po_number = request.POST.get('po_number')
        equipment.fund_source = request.POST.get('fund_source')
        equipment.supplier = request.POST.get('supplier')
        equipment.item_amount = request.POST.get('item_amount')
        equipment.assigned_to = request.POST.get('assigned_to') or None
        equipment.location = request.POST.get('location')
        equipment.current_location = request.POST.get('current_location')
        equipment.project_name = request.POST.get('project_name') or None
        equipment.end_user = request.POST.get('end_user') or None
        
        # Handle reason fields
        equipment.assignment_reason = request.POST.get('assignment_reason') or None
        equipment.user_change_reason = request.POST.get('user_change_reason') or None
        equipment.location_change_reason = request.POST.get('location_change_reason') or None
        equipment.current_location_reason = request.POST.get('current_location_reason') or None
        
        equipment.emp_id = request.POST.get('emp_id')
        equipment.category_id = request.POST.get('category_id')
        equipment.status_id = request.POST.get('status_id')
        
        # Handle damage reason - only save if status is "Damaged"
        # Handle lost remarks - only save if status is "Lost"
        status_id = request.POST.get('status_id')
        if status_id:
            try:
                selected_status = Status.objects.get(id=status_id)
                if selected_status.name == 'Damaged':
                    equipment.damage_reason = request.POST.get('damage_reason', '').strip()
                    equipment.lost_remarks = None  # Clear lost remarks if not lost
                elif selected_status.name == 'Lost':
                    equipment.lost_remarks = request.POST.get('lost_remarks', '').strip()
                    equipment.damage_reason = None  # Clear damage reason if not damaged
                else:
                    equipment.damage_reason = None  # Clear if neither damaged nor lost
                    equipment.lost_remarks = None
            except Status.DoesNotExist:
                equipment.damage_reason = None
                equipment.lost_remarks = None
        
        equipment.updated_by = request.user

        if request.FILES.get('user_image'):
            equipment.user_image = request.FILES['user_image']

        if request.FILES.get('order_receipt'):
            equipment.order_receipt = request.FILES['order_receipt']

        # Save equipment first
        equipment.save()
        
        # Handle replacement documents if status is "Lost"
        replacement_files = request.FILES.getlist('replacement_documents')
        if replacement_files and status_id:
            try:
                selected_status = Status.objects.get(id=status_id)
                if selected_status.name == 'Lost':
                    for file in replacement_files:
                        ReplacementDocument.objects.create(
                            equipment=equipment,
                            document=file,
                            original_filename=file.name,
                            uploaded_by=request.user
                        )
            except Status.DoesNotExist:
                pass
        
        EquipmentActionLog.objects.create(
            equipment=equipment,
            action='edit',
            user=request.user,
            summary=f"Edited equipment: {equipment.item_name} (Property #: {equipment.item_propertynum})"
        )

        field_labels = {
            'item_propertynum': 'Property #',
            'item_name': 'Name',
            'item_desc': 'Description',
            'item_purdate': 'Purchase Date',
            'po_number': 'PO Number',
            'fund_source': 'Fund Source',
            'supplier': 'Supplier',
            'item_amount': 'Amount',
            'assigned_to': 'Assigned To',
            'location': 'Deployment Location',
            'current_location': 'Current Location',
            'project_name': 'Project Name',
            'end_user': 'End User',
            'emp_id': 'Employee',
            'category_id': 'Category',
            'status_id': 'Status',
            'damage_reason': 'Damage Reason',
            'lost_remarks': 'Lost Remarks',
            # Reason field labels removed since they're not tracked in history
            # 'assignment_reason': 'Assignment Change Reason',
            # 'user_change_reason': 'End User Change Reason',
            # 'location_change_reason': 'Location Change Reason',
            # 'current_location_reason': 'Current Location Change Reason',
        }
        for field, old in original.items():
            new = getattr(equipment, field)
            
            # Skip if values are the same
            if old == new:
                continue
                
            # Initialize variables
            old_val = ''
            new_val = ''
            reason = None
            
            # Get the appropriate reason for specific fields
            if field == 'assigned_to':
                reason = equipment.assignment_reason
            elif field == 'end_user':
                reason = equipment.user_change_reason
            elif field == 'location':
                reason = equipment.location_change_reason
            elif field == 'current_location':
                reason = equipment.current_location_reason
            
            # Handle different field types
            if field == 'category_id':
                old_val = str(Category.objects.get(pk=old).name) if old else ''
                new_val = str(equipment.category.name) if equipment.category else ''
            elif field == 'status_id':
                old_val = str(Status.objects.get(pk=old).name) if old else ''
                new_val = str(equipment.status.name) if equipment.status else ''
            elif field == 'emp_id':
                old_val = str(User.objects.get(pk=old).get_full_name()) if old else ''
                new_val = str(equipment.emp.get_full_name()) if equipment.emp else ''
            elif field == 'item_purdate':
                old_val = old.strftime('%Y-%m-%d') if old else ''
                new_val = new.strftime('%Y-%m-%d') if new else ''
            elif field == 'item_amount':
                try:
                    old_amount = float(old) if old not in (None, '', 'None') else 0.0
                except Exception:
                    old_amount = 0.0
                try:
                    new_amount = float(new) if new not in (None, '', 'None') else 0.0
                except Exception:
                    new_amount = 0.0
                old_val = f"₱{old_amount:,.2f}" if old not in ('', None, 'None') else ''
                new_val = f"₱{new_amount:,.2f}" if new not in ('', None, 'None') else ''
                # Skip if amounts are the same
                if old_amount == new_amount:
                    continue
            else:
                # Normalize for text fields: treat None, '', and 'None' as equivalent, and strip whitespace
                def norm(val):
                    if val is None or val == '' or str(val).strip().lower() == 'none':
                        return ''
                    return str(val).strip()
                old_val = norm(old)
                new_val = norm(new)
                # Skip if normalized values are the same
                if old_val == new_val:
                    continue
            
            # Create history record if values are different
            if old_val != new_val:
                EquipmentHistory.objects.create(
                    equipment=equipment,
                    field_changed=field_labels.get(field, field),
                    old_value=old_val,
                    new_value=new_val,
                    reason=reason,
                    action='Edited',
                    changed_by=request.user
                )

        return redirect('equipments:index')

    # Get existing values for dropdowns
    existing_end_users = Equipment.objects.exclude(end_user__isnull=True).exclude(end_user__exact='').values_list('end_user', flat=True).distinct().order_by('end_user')
    existing_assigned_to = Equipment.objects.exclude(assigned_to__isnull=True).exclude(assigned_to__exact='').values_list('assigned_to', flat=True).distinct().order_by('assigned_to')
    existing_project_names = Equipment.objects.exclude(project_name__isnull=True).exclude(project_name__exact='').values_list('project_name', flat=True).distinct().order_by('project_name')
    existing_locations = Equipment.objects.exclude(location__isnull=True).exclude(location__exact='').values_list('location', flat=True).distinct().order_by('location')
    existing_current_locations = Equipment.objects.exclude(current_location__isnull=True).exclude(current_location__exact='').values_list('current_location', flat=True).distinct().order_by('current_location')

    return render(request, 'equipments/edit.html', {
        'equipment': equipment,
        'categories': categories,
        'statuses': statuses,
        'users': users,
        'existing_end_users': existing_end_users,
        'existing_assigned_to': existing_assigned_to,
        'existing_project_names': existing_project_names,
        'existing_locations': existing_locations,
        'existing_current_locations': existing_current_locations,
    })

@login_required
@user_passes_test(is_admin_or_superadmin)
def delete_equipment(request, id):
    equipment = get_object_or_404(Equipment, id=id)
    item_name = equipment.item_name
    item_propertynum = equipment.item_propertynum
    
    # Check if equipment was returned (for redirect purposes)
    was_returned = equipment.is_returned
    
    # Log the delete action before deleting equipment and logs
    delete_log = EquipmentActionLog.objects.create(
        equipment=None,  # Set to None to avoid cascade issues
        action='delete',
        user=request.user,
        summary=f"Deleted equipment: {item_name} (Property #: {item_propertynum})"
    )
    
    # Delete related action logs and return documents (except the delete log we just created)
    EquipmentActionLog.objects.filter(equipment_id=equipment.id).exclude(id=delete_log.id).delete()
    ReturnDocument.objects.filter(equipment_id=equipment.id).delete()
    equipment.delete()
    
    messages.success(request, f"Equipment '{item_name}' has been permanently deleted.")
    
    # Redirect back to appropriate page based on where the user came from
    referer = request.META.get('HTTP_REFERER', '')
    if 'returned' in referer or was_returned:
        return redirect('equipments:returned')
    else:
        return redirect('equipments:index')


@login_required
@secretariat_required
def dashboard(request):
    total_equipments = Equipment.objects.filter(is_returned=False).count()
    total_archived = Equipment.objects.filter(is_archived=True).count()
    total_returned = Equipment.objects.filter(is_returned=True).count()

    # Status counts for pie chart (only for active equipment)
    status_counts = Equipment.objects.filter(is_archived=False, is_returned=False).values('status__name').annotate(
        name=F('status__name'), count=Count('id')
    )
    status_labels = [s['name'] for s in status_counts]
    status_data = [s['count'] for s in status_counts]

    # Category counts for bar chart (only for active equipment)
    from django.db.models import Q
    categories = Category.objects.annotate(
        count=Count('equipment', filter=Q(equipment__is_archived=False, equipment__is_returned=False))
    )
    category_labels = [cat.name for cat in categories]
    category_counts = [cat.count for cat in categories]

    # Recent equipments (only active ones)
    recent_equipments = Equipment.objects.filter(is_archived=False, is_returned=False).order_by('-id')[:5]

    # Equipments acquired per year (from PO date) - only active equipment
    from django.db.models.functions import ExtractYear
    year_qs = Equipment.objects.filter(is_archived=False, is_returned=False).exclude(item_purdate=None).values(year=ExtractYear('item_purdate')).annotate(count=Count('id')).order_by('year')
    year_labels = [str(x['year']) for x in year_qs]
    year_data = [x['count'] for x in year_qs]

    # Total number and total cost of equipment per end user (currently held)
    enduser_qs = Equipment.objects.filter(is_archived=False, is_returned=False).exclude(end_user__isnull=True).exclude(end_user='').values('end_user').annotate(
        count=Count('id'),
        total=Sum('item_amount')
    ).order_by('-count')
    enduser_labels = [x['end_user'] for x in enduser_qs]
    enduser_counts = [x['count'] for x in enduser_qs]
    enduser_amounts = [float(x['total'] or 0) for x in enduser_qs]

    # Equipments by Assigned To: Count and Total Cost
    assigned_qs = Equipment.objects.filter(is_archived=False, is_returned=False).exclude(assigned_to__isnull=True).exclude(assigned_to='').values('assigned_to').annotate(
        count=Count('id'),
        total=Sum('item_amount')
    ).order_by('-count')
    assigned_labels = [x['assigned_to'] for x in assigned_qs]
    assigned_counts = [x['count'] for x in assigned_qs]
    assigned_amounts = [float(x['total'] or 0) for x in assigned_qs]

    # Equipments by Item Name: Count (only active equipment)
    name_qs = Equipment.objects.filter(is_archived=False, is_returned=False).values('item_name').annotate(count=Count('id')).order_by('-count')
    itemname_labels = [x['item_name'] for x in name_qs]
    itemname_counts = [x['count'] for x in name_qs]

    # Equipments by Project Name: Count (only active equipment)
    project_qs = Equipment.objects.filter(is_archived=False, is_returned=False).exclude(project_name__isnull=True).exclude(project_name='').values('project_name').annotate(count=Count('id')).order_by('-count')
    project_labels = [x['project_name'] for x in project_qs]
    project_counts = [x['count'] for x in project_qs]
    
    context = {
        'total_equipments': total_equipments,
        'total_archived': total_archived,
        'total_returned': total_returned,
        'status_counts': status_counts,
        'status_labels': status_labels,
        'status_data': status_data,
        'category_labels': json.dumps(category_labels),
        'category_counts': json.dumps(category_counts),
        'recent_equipments': recent_equipments,
        'month_labels': json.dumps(year_labels),  # Used by the chart, but now years
        'month_data': json.dumps(year_data),      # Used by the chart, but now yearly counts
        'enduser_labels': json.dumps(enduser_labels),
        'enduser_counts': json.dumps(enduser_counts),
        'enduser_amounts': json.dumps(enduser_amounts),
        'assigned_labels': json.dumps(assigned_labels),
        'assigned_counts': json.dumps(assigned_counts),
        'assigned_amounts': json.dumps(assigned_amounts),
        'itemname_labels': json.dumps(itemname_labels),
        'itemname_counts': json.dumps(itemname_counts),
        'project_labels': json.dumps(project_labels),
        'project_counts': json.dumps(project_counts),
        'is_admin': is_admin(request.user),
        'is_encoder': is_encoder(request.user),
        'is_superadmin': is_superadmin(request.user),
        'is_client': is_client(request.user),
    }
    return render(request, 'equipments/dashboard.html', context)

# User management views - REMOVED since WEIMS uses main WESMAARRDEC user system
# The main website already handles user management, so these are redundant

# @login_required
# @user_passes_test(is_admin_or_superadmin)
# def user(request):
#     users = User.objects.all().order_by('-date_joined')
#     return render(request, 'equipments/user.html', {
#         'users': users,
#         'is_admin': is_admin(request.user),
#         'is_superadmin': is_superadmin(request.user),
#         'is_encoder': is_encoder(request.user),
#     })

# @login_required
# @user_passes_test(is_admin_or_superadmin)
# def add_user(request):
#     error = None
#     if request.method == 'POST':
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         email = request.POST.get('email')
#         role = request.POST.get('role')
#         first_name = request.POST.get('first_name')
#         last_name = request.POST.get('last_name')
#
#         # Check for required fields
#         if not all([username, password, role, first_name, last_name]):
#             error = "All fields are required."
#         elif User.objects.filter(username=username).exists():
#             error = "Username already exists."
#         else:
#             user = User.objects.create_user(
#                 username=username,
#                 password=password,
#                 email=email,
#                 first_name=first_name,
#                 last_name=last_name
#             )
#             group = Group.objects.get(name=role)
#             user.groups.add(group)
#             return redirect('equipments:user')
#             
#     return render(request, 'equipments/add_user.html', {'error': error})

# @login_required
# @user_passes_test(is_admin_or_superadmin)
# def edit_user(request, user_id):
#     user = get_object_or_404(User, id=user_id)
#     if request.method == 'POST':
#         user.username = request.POST.get('username', user.username)
#         user.email = request.POST.get('email', user.email)
#         user.first_name = request.POST.get('first_name', user.first_name)
#         user.last_name = request.POST.get('last_name', user.last_name)
#         role = request.POST.get('role')
#         if role:
#             # Remove from all groups and add to the selected one
#             user.groups.clear()
#             group = Group.objects.get(name=role)
#             user.groups.add(group)
#         user.save()
#         return redirect('equipments:user')
#     return render(request, 'equipments/edit_user.html', {'user_obj': user})

# @login_required
# @user_passes_test(is_admin_or_superadmin)
# def delete_user(request, user_id):
#     user = get_object_or_404(User, id=user_id)
#     if request.method == 'POST':
#         user.delete()
#         return redirect('equipments:user')
#     return render(request, 'equipments/confirm_delete_user.html', {'user_obj': user})

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def category_list(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            if Category.objects.filter(name__iexact=name).exists():
                messages.warning(request, "Category already exists.")
            else:
                Category.objects.create(name=name)
                EquipmentActionLog.objects.create(
                    equipment=None,
                    action='create',
                    user=request.user,
                    summary=f"Created new category: '{name}'"
                )
                messages.success(request, "Category added successfully.")
        else:
            messages.error(request, "Category name cannot be empty.")
        return redirect('equipments:category')  

    categories = Category.objects.all().order_by('name')
    return render(request, 'equipments/category.html', {
        'categories': categories, 
        'is_admin': is_admin(request.user),})

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def edit_category(request, id):
    category = get_object_or_404(Category, id=id)
    if request.method == 'POST':
        old_name = category.name
        name = request.POST.get('name', '').strip()
        if name:
            category.name = name
            category.save()
            EquipmentActionLog.objects.create(
                equipment=None,
                action='edit',
                user=request.user,
                summary=f"Updated category: '{old_name}' → '{name}'"
            )
            messages.success(request, "Category updated successfully.")
        else:
            messages.error(request, "Category name cannot be empty.")
    return redirect('equipments:category')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def delete_category(request, id):
    category = get_object_or_404(Category, id=id)
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        EquipmentActionLog.objects.create(
            equipment=None,
            action='delete',
            user=request.user,
            summary=f"Deleted category: '{category_name}'"
        )
        messages.success(request, "Category deleted successfully.")
    return redirect('equipments:category')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def status_list(request):
    statuses = Status.objects.all()
    return render(request, 'equipments/status.html', {
        'statuses': statuses
        , 'is_admin': is_admin(request.user),})

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def add_status(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            if Status.objects.filter(name__iexact=name).exists():
                messages.warning(request, "Status already exists.")
            else:
                Status.objects.create(name=name)
                EquipmentActionLog.objects.create(
                    equipment=None,
                    action='create',
                    user=request.user,
                    summary=f"Created new status: '{name}'"
                )
                messages.success(request, "Status added successfully.")
        else:
            messages.error(request, "Status name cannot be empty.")
    return redirect('equipments:status')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def edit_status(request, id):
    status = get_object_or_404(Status, id=id)
    if request.method == 'POST':
        old_name = status.name
        name = request.POST.get('name', '').strip()
        if name:
            status.name = name
            status.save()
            EquipmentActionLog.objects.create(
                equipment=None,
                action='edit',
                user=request.user,
                summary=f"Updated status: '{old_name}' → '{name}'"
            )
            messages.success(request, "Status updated successfully.")
        else:
            messages.error(request, "Status name cannot be empty.")
    return redirect('equipments:status')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def delete_status(request, id):
    status = get_object_or_404(Status, id=id)
    if request.method == 'POST':
        status_name = status.name
        status.delete()
        EquipmentActionLog.objects.create(
            equipment=None,
            action='delete',
            user=request.user,
            summary=f"Deleted status: '{status_name}'"
        )
        messages.success(request, "Status deleted successfully.")
    return redirect('equipments:status')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def export_excel(request):
    """Export equipment data to Excel format matching the import template"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipment Data"
        
        # Headers matching import format (14 columns)
        headers = [
            'Property Number', 'Item Name', 'Item Description', 'Additional Info', 
            'Purchase Date', 'PO Number', 'Fund Source', 'Supplier', 
            'Amount', 'Project Name', 'Assigned To', 'End User', 
            'Location', 'Current Location'
        ]
        
        # Style headers
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Write data
        for row_idx, equipment in enumerate(Equipment.objects.all(), 2):
            data = [
                equipment.item_propertynum or '',
                equipment.item_name or '',
                equipment.item_desc or '',
                equipment.additional_info or '',
                equipment.item_purdate.strftime('%Y-%m-%d') if equipment.item_purdate else '',
                equipment.po_number or '',
                equipment.fund_source or '',
                equipment.supplier or '',
                float(equipment.item_amount) if equipment.item_amount else 0.00,
                equipment.project_name or '',
                equipment.assigned_to or '',
                equipment.end_user or '',
                equipment.location or '',
                equipment.current_location or ''
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="equipment_template.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, "openpyxl library not installed. Please install it to export Excel files.")
        return redirect('equipments:index')
    except Exception as e:
        messages.error(request, f"Error creating Excel file: {str(e)}")
        return redirect('equipments:index')

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            imported_count = 0
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    cleaned_row = [cell if cell not in ('', None) else None for cell in row]
                    while len(cleaned_row) < 14:
                        cleaned_row.append(None)
                    
                    propertynum = cleaned_row[0]
                    if propertynum and Equipment.objects.filter(item_propertynum=propertynum).exists():
                        continue
                    
                    def get_safe_value(index, default=None):
                        return cleaned_row[index] if index < len(cleaned_row) else default
                    
                    Equipment.objects.create(
                        item_propertynum=get_safe_value(0),
                        item_name=get_safe_value(1) or "Imported Item",
                        item_desc=get_safe_value(2),
                        additional_info=(get_safe_value(3)[:300] if get_safe_value(3) else None),
                        item_purdate=parse_date(get_safe_value(4)),
                        po_number=get_safe_value(5),
                        fund_source=get_safe_value(6),
                        supplier=get_safe_value(7),
                        item_amount=get_safe_value(8) or 0.00,
                        project_name=get_safe_value(9),
                        assigned_to=get_safe_value(10),
                        end_user=get_safe_value(11),
                        location=get_safe_value(12),
                        current_location=get_safe_value(13),
                        category=Category.objects.get(pk=1),  
                        status=Status.objects.get(pk=1),      
                        emp=request.user,
                        created_by=request.user,
                        updated_by=request.user,
                    )
                    imported_count += 1
                    
                    EquipmentActionLog.objects.create(
                        equipment=Equipment.objects.filter(item_propertynum=propertynum).first() if propertynum else None,
                        action='create',
                        user=request.user,
                        summary=f"Imported equipment from Excel: {get_safe_value(1) or 'Unknown'} (Property #: {propertynum or 'None'})"
                    )
                    
                except Category.DoesNotExist:
                    request.session['import_error'] = {
                        'title': 'Category Not Found',
                        'message': f'Import stopped at Row {idx}: The specified category does not exist in the system.',
                        'details': 'Please ensure all equipment categories are created before importing. Navigate to <strong>Admin Panel → Equipment Categories</strong> to add missing categories.'
                    }
                    return redirect('equipments:index')
                    
                except Status.DoesNotExist:
                    request.session['import_error'] = {
                        'title': 'Status Not Found',
                        'message': f'Import stopped at Row {idx}: The specified status does not exist in the system.',
                        'details': 'Please ensure all equipment statuses are created before importing. Navigate to <strong>Admin Panel → Equipment Statuses</strong> to add missing statuses.'
                    }
                    return redirect('equipments:index')
                    
                except Exception as e:
                    error_title = "Import Failed"
                    error_msg = f"Import stopped at Row {idx}"
                    error_details = str(e)
                    
                    if "Category matching query does not exist" in str(e):
                        error_title = "Category Not Found"
                        error_details = "The specified category does not exist in the system. Please verify that all categories referenced in your Excel file have been created. Navigate to <strong>Admin Panel → Equipment Categories</strong> to manage categories."
                    elif "Status matching query does not exist" in str(e):
                        error_title = "Status Not Found"
                        error_details = "The specified status does not exist in the system. Please verify that all statuses referenced in your Excel file have been created. Navigate to <strong>Admin Panel → Equipment Statuses</strong> to manage statuses."
                    elif "list index out of range" in str(e):
                        error_title = "Invalid File Format"
                        error_details = f"Insufficient data columns detected. Expected 14 columns, but found {len(row) if 'row' in locals() else 'unknown'}. Please ensure your Excel file follows the correct template format."
                    elif "does not exist" in str(e).lower():
                        error_title = "Missing Reference Data"
                        error_details = "Referenced data not found in the system. Please ensure all categories, statuses, and related data are properly configured before importing."
                    
                    request.session['import_error'] = {
                        'title': error_title,
                        'message': error_msg,
                        'details': error_details
                    }
                    return redirect('equipments:index')
            
            messages.success(request, f"Excel import completed successfully. {imported_count} equipment(s) imported.")
            
        except Exception as e:
            request.session['import_error'] = {
                'title': 'File Processing Error',
                'message': 'Unable to process the Excel file.',
                'details': f"Error: {str(e)}. Please ensure you're uploading a valid Excel file (.xlsx or .xls)."
            }
            return redirect('equipments:index')
    else:
        messages.error(request, "Please select an Excel file to import.")
    
    return redirect('equipments:index')

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except Exception:
        return None  # or raise

@require_POST
@login_required
@user_passes_test(is_admin_superadmin_encoder)
def bulk_update_equipment(request):
    
    ids = request.POST.get('equipment_ids', '')
    status_id = request.POST.get('status_id')
    category_id = request.POST.get('category_id')
    if not ids:
        return JsonResponse({'error': 'No equipment selected.'}, status=400)
    
    # Validate status - prevent Lost and Damaged from bulk updates
    if status_id:
        try:
            status = Status.objects.get(id=status_id)
            if status.name in ['Lost', 'Damaged']:
                return JsonResponse({
                    'error': f'Cannot bulk update to "{status.name}" status. This status requires additional information that must be set individually for each equipment.'
                }, status=400)
        except Status.DoesNotExist:
            return JsonResponse({'error': 'Invalid status selected.'}, status=400)
    
    id_list = [int(i) for i in ids.split(',') if i.isdigit()]
    qs = Equipment.objects.filter(id__in=id_list)
    updates = {}
    if status_id:
        updates['status_id'] = status_id
    if category_id:
        updates['category_id'] = category_id
    if updates:
        # Log the bulk update action
        equipment_names = list(qs.values_list('item_name', 'item_propertynum'))
        update_details = []
        if status_id:
            try:
                status_name = Status.objects.get(id=status_id).name
                update_details.append(f"Status to '{status_name}'")
            except Status.DoesNotExist:
                update_details.append(f"Status to ID {status_id}")
        if category_id:
            try:
                category_name = Category.objects.get(id=category_id).name
                update_details.append(f"Category to '{category_name}'")
            except Category.DoesNotExist:
                update_details.append(f"Category to ID {category_id}")
        
        qs.update(**updates)
        
        # Create a single log entry for the bulk action
        equipment_list = ', '.join([f"{name} (#{prop})" for name, prop in equipment_names[:5]])
        if len(equipment_names) > 5:
            equipment_list += f" and {len(equipment_names) - 5} more"
        
        EquipmentActionLog.objects.create(
            equipment=None,  # Bulk action, no single equipment
            action='edit',
            user=request.user,
            summary=f"Bulk updated {len(equipment_names)} equipment(s): {equipment_list} - Changed: {', '.join(update_details)}"
        )
    return JsonResponse({'success': True})

@login_required
@role_required_with_feedback(is_admin_superadmin_encoder)
def bulk_archive_equipment(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed.'}, status=405)
    
    ids = request.POST.get('equipment_ids', '')
    if not ids:
        return JsonResponse({'error': 'No equipment selected.'}, status=400)
    
    id_list = [int(i) for i in ids.split(',') if i.isdigit()]
    qs = Equipment.objects.filter(id__in=id_list, is_archived=False)  # Only archive non-archived equipment
    
    if not qs.exists():
        return JsonResponse({'error': 'No valid equipment found to archive.'}, status=400)
    
    # Get equipment details for logging before archiving
    equipment_names = list(qs.values_list('item_name', 'item_propertynum'))
    archived_count = qs.count()
    
    # Archive the equipment
    qs.update(is_archived=True)
    
    # Create history log entries for each equipment
    for equipment in qs:
        EquipmentHistory.objects.create(
            equipment=equipment,
            field_changed='is_archived',
            old_value='False',
            new_value='True',
            changed_by=request.user,
            change_reason='Bulk archive operation'
        )
    
    # Create a single action log entry for the bulk operation
    equipment_list = ', '.join([f"{name} (#{prop})" for name, prop in equipment_names[:5]])
    if len(equipment_names) > 5:
        equipment_list += f" and {len(equipment_names) - 5} more"
    
    EquipmentActionLog.objects.create(
        equipment=None,  # Bulk action, no single equipment
        action='archive',
        user=request.user,
        summary=f"Bulk archived {archived_count} equipment(s): {equipment_list}"
    )
    
    return JsonResponse({
        'success': True, 
        'archived_count': archived_count,
        'message': f'Successfully archived {archived_count} equipment(s).'
    })

@login_required
@role_required_with_feedback(is_admin_or_superadmin)  # Only admin and superadmin can bulk delete
def bulk_delete_equipment(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed.'}, status=405)
    
    ids = request.POST.get('equipment_ids', '')
    if not ids:
        return JsonResponse({'error': 'No equipment selected.'}, status=400)
    
    id_list = [int(i) for i in ids.split(',') if i.isdigit()]
    qs = Equipment.objects.filter(id__in=id_list)
    
    if not qs.exists():
        return JsonResponse({'error': 'No valid equipment found to delete.'}, status=400)
    
    # Get equipment details for logging before deletion
    equipment_details = []
    for equipment in qs:
        equipment_details.append({
            'id': equipment.id,
            'name': equipment.item_name or 'Unknown',
            'property_num': equipment.item_propertynum or 'N/A',
            'category': equipment.category.name if equipment.category else 'None',
            'status': equipment.status.name if equipment.status else 'None',
            'end_user': equipment.end_user or 'None',
            'assigned_to': equipment.assigned_to or 'None'
        })
    
    deleted_count = len(equipment_details)
    
    # Create history logs before deletion (since we can't log after deletion)
    for eq_detail in equipment_details:
        # Create a final action log for each equipment being deleted
        EquipmentActionLog.objects.create(
            equipment=None,  # Equipment will be deleted, so no reference
            action='delete',
            user=request.user,
            summary=f"Equipment PERMANENTLY DELETED - ID: {eq_detail['id']}, Name: {eq_detail['name']}, Property #: {eq_detail['property_num']}, Category: {eq_detail['category']}, Status: {eq_detail['status']}, End User: {eq_detail['end_user']}, Assigned To: {eq_detail['assigned_to']}"
        )
    
    # Create a single bulk action log entry
    equipment_list = ', '.join([f"{eq['name']} (#{eq['property_num']})" for eq in equipment_details[:5]])
    if len(equipment_details) > 5:
        equipment_list += f" and {len(equipment_details) - 5} more"
    
    EquipmentActionLog.objects.create(
        equipment=None,  # Bulk action, no single equipment
        action='bulk_delete',
        user=request.user,
        summary=f"BULK HARD DELETE - Permanently deleted {deleted_count} equipment(s): {equipment_list}"
    )
    
    # Now perform the actual deletion
    qs.delete()
    
    return JsonResponse({
        'success': True, 
        'deleted_count': deleted_count,
        'message': f'Successfully deleted {deleted_count} equipment(s) permanently.'
    })

@login_required
@role_required_with_feedback(is_viewer_or_above)
def returned(request):
    equipments = Equipment.objects.filter(is_returned=True)
    total_returned = equipments.count()
    return render(request, 'equipments/returned.html', {
        'equipments': equipments,
        'total_returned': total_returned,
        'is_admin': is_admin(request.user),
        'is_encoder': is_encoder(request.user),
        'is_superadmin': is_superadmin(request.user),
    })

@login_required
@role_required_with_feedback(is_viewer_or_above)
def returned_equipment_table_json(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        qs = Equipment.objects.filter(is_returned=True).select_related('category', 'status', 'emp')

        if search_value:
            qs = qs.filter(
                Q(item_propertynum__icontains=search_value) |
                Q(item_name__icontains=search_value) |
                Q(item_desc__icontains=search_value) |
                Q(category__name__icontains=search_value) |
                Q(status__name__icontains=search_value) |
                Q(po_number__icontains=search_value) |
                Q(fund_source__icontains=search_value) |
                Q(supplier__icontains=search_value) |
                Q(item_amount__icontains=search_value) |
                Q(assigned_to__icontains=search_value) |
                Q(end_user__icontains=search_value) |
                Q(location__icontains=search_value) |
                Q(current_location__icontains=search_value) |
                Q(item_purdate__icontains=search_value) |
                Q(project_name__icontains=search_value)
            )

        total = Equipment.objects.filter(is_returned=True).count()
        filtered = qs.count()
        equipments = qs.order_by('-updated_at')[start:start+length]

        data = []
        for eq in equipments:
            # Get all return documents for this equipment
            return_docs = ReturnDocument.objects.filter(equipment=eq).order_by('-uploaded_at')
            
            # Create document data structure
            documents_data = []
            if return_docs.exists():
                for doc in return_docs:
                    documents_data.append({
                        'url': doc.document.url,
                        'filename': doc.original_filename or doc.document.name.split('/')[-1],
                        'uploaded_at': doc.uploaded_at.strftime("%b %d, %Y %H:%M")
                    })
            elif eq.return_document:  # Fallback to old single document field
                documents_data.append({
                    'url': eq.return_document.url,
                    'filename': eq.return_document.name.split('/')[-1],
                    'uploaded_at': 'Legacy'
                })
            
            # Action buttons for admin/superadmin only
            actions = ''
            if is_admin(request.user) or is_superadmin(request.user):
                actions = f'''
                <div class="dropdown">
                    <button class="btn btn-outline-secondary btn-sm dropdown-toggle" type="button" data-bs-toggle="dropdown" aria-expanded="false">
                        Actions
                    </button>
                    <ul class="dropdown-menu">
                        <li>
                            <a class="dropdown-item text-success" href="/weims/reissue/{eq.id}/" onclick="return confirm('Reissue this equipment? It will be moved back to active inventory.');">
                                <i class="bi bi-arrow-repeat"></i> Reissue
                            </a>
                        </li>
                        <li>
                            <a class="dropdown-item" href="/weims/edit/{eq.id}/">
                                <i class="bi bi-pencil-square"></i> Edit
                            </a>
                        </li>
                        <li><hr class="dropdown-divider"></li>
                        <li>
                            <a class="dropdown-item text-danger" href="/weims/delete/{eq.id}/" onclick="return confirm('Are you sure you want to permanently delete this equipment? This action cannot be undone.');">
                                <i class="bi bi-trash"></i> Delete
                            </a>
                        </li>
                    </ul>
                </div>
                '''
            
            data.append([
                eq.id,
                f'<img src="{eq.user_image.url if eq.user_image else ""}" class="img-thumbnail" style="width:32px;height:32px;object-fit:cover;">',
                eq.item_propertynum,
                eq.item_name,
                eq.item_desc or 'None',
                eq.returned_by or 'None',
                json.dumps(documents_data),  # Pass documents as JSON
                eq.updated_at.strftime("%b %d, %Y") if eq.updated_at else 'None',
                eq.return_remarks or 'None',
                eq.return_condition or 'None',
                eq.return_type or 'None',
                eq.received_by or 'None',
                actions  # Add actions column
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total,
            'recordsFiltered': filtered,
            'data': data,
        })
        
    except Exception as e:
        import traceback
        # Return error details for debugging
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'draw': request.GET.get('draw', 1),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        }, status=500)

@require_POST
@login_required
@role_required_with_feedback(is_viewer_or_above)
def return_equipment(request):
    eq_id = request.POST.get('equipment_id')
    files = request.FILES.getlist('return_document')  # Get multiple files
    remarks = request.POST.get('return_remarks')
    condition = request.POST.get('return_condition')
    return_type = request.POST.get('return_type')
    returned_by = request.POST.get('returned_by') 
    received_by = request.POST.get('received_by')
    
    if not eq_id or not files:
        messages.error(request, "Equipment and at least one document are required.")
        return redirect('equipments:index')
    
    eq = get_object_or_404(Equipment, id=eq_id)
    eq.is_returned = True
    
    # Keep the first file in the original field for backward compatibility
    if files:
        eq.return_document = files[0]
    
    eq.return_remarks = remarks
    eq.return_condition = condition
    eq.return_type = return_type
    eq.returned_by = returned_by
    eq.received_by = received_by
    eq.save()
    
    # Save all files to the new ReturnDocument model
    for file in files:
        ReturnDocument.objects.create(
            equipment=eq,
            document=file,
            original_filename=file.name,
            uploaded_by=request.user
        )
    
    EquipmentActionLog.objects.create(
        equipment=eq,
        action='return',
        user=request.user,
        summary=f"Returned equipment: {eq.item_name} (Property #: {eq.item_propertynum}) with {len(files)} document(s)"
    )
    
    messages.success(request, f"Equipment marked as returned with {len(files)} document(s).")
    return redirect('equipments:index')

@login_required
@role_required_with_feedback(is_viewer_or_above)
def archived_equipments(request):
    context = {
        'is_admin': is_admin(request.user),
        'is_encoder': is_encoder(request.user),
        'is_client': is_client(request.user),
        'is_superadmin': is_superadmin(request.user),
    }
    return render(request, 'equipments/archived_list.html', context)

@login_required
@role_required_with_feedback(is_viewer_or_above)
def archive_equipment(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    
    # Check if the request is coming from returned equipment page
    referer = request.META.get('HTTP_REFERER', '')
    if 'returned' in referer or equipment.is_returned:
        messages.error(request, "Cannot archive returned equipment from this page. Please use the delete option instead.")
        return redirect('equipments:returned')
    
    equipment.is_archived = True
    equipment.date_archived = timezone.now()
    equipment.archived_by = request.user
    equipment.save()
    EquipmentActionLog.objects.create(
        equipment=equipment,
        action='archive',
        user=request.user,
        summary=f"Archived equipment: {equipment.item_name} (Property #: {equipment.item_propertynum})"
    )
    messages.success(request, "Equipment sent to archive.")
    
    # Redirect back to appropriate page based on where the user came from
    return redirect('equipments:index')

@login_required
@role_required_with_feedback(is_viewer_or_above)
def archived_equipment_table_json(request):
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')

        qs = Equipment.objects.filter(is_archived=True).select_related('category', 'status', 'emp')

        if search_value:
            qs = qs.filter(
                Q(item_propertynum__icontains=search_value) |
                Q(item_name__icontains=search_value) |
                Q(item_desc__icontains=search_value) |
                Q(category__name__icontains=search_value) |
                Q(status__name__icontains=search_value) |
                Q(po_number__icontains=search_value) |
                Q(fund_source__icontains=search_value) |
                Q(supplier__icontains=search_value) |
                Q(item_amount__icontains=search_value) |
                Q(assigned_to__icontains=search_value) |
                Q(end_user__icontains=search_value) |
                Q(location__icontains=search_value) |
                Q(current_location__icontains=search_value) |
                Q(item_purdate__icontains=search_value) |
                Q(project_name__icontains=search_value)
            )

        total = Equipment.objects.filter(is_archived=True).count()
        filtered = qs.count()
        equipments = qs.order_by('-item_purdate')[start:start+length]

        data = []
        for eq in equipments:
            # Only show the Recover (Unarchive) button if user is admin or superadmin
            actions = ''
            if is_admin(request.user) or is_superadmin(request.user):
                actions = f'''
                <a class="btn btn-sm btn-outline-secondary" href="/weims/unarchive/{eq.id}/" onclick="return confirm('Unarchive this equipment?');">
                  <i class="bi bi-arrow-counterclockwise"></i> Recover
                </a>
                '''
            data.append([
                eq.id,
                f'<img src="{eq.user_image.url if eq.user_image else ""}" class="img-thumbnail" style="width:32px;height:32px;object-fit:cover;">',
                eq.item_propertynum,
                eq.item_name,
                eq.item_desc or 'None',
                eq.po_number or 'None',
                f'₱{eq.item_amount:,.2f}',
                eq.end_user or 'None',
                eq.category.name,
                eq.status.name + ("<span class='badge bg-secondary ms-1'>Deleted</span>" if eq.is_archived else ""),
                eq.date_archived.strftime('%Y-%m-%d %H:%M') if eq.date_archived else 'None',
                f'{eq.archived_by.get_full_name() if eq.archived_by else "None"}',
                actions
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total,
            'recordsFiltered': filtered,
            'data': data,
        })
        
    except Exception as e:
        import traceback
        # Return error details for debugging
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'draw': request.GET.get('draw', 1),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': []
        }, status=500)

@login_required
@user_passes_test(is_admin_superadmin_encoder)
def unarchive_equipment(request, pk):
    eq = get_object_or_404(Equipment, pk=pk)
    eq.is_archived = False
    eq.save()
    EquipmentActionLog.objects.create(
    equipment=eq,
    action='unarchive',
    user=request.user,
    summary=f"Unarchived equipment: {eq.item_name} (Property #: {eq.item_propertynum})"
)
    return redirect('equipments:archived_equipments')

@login_required
@user_passes_test(is_admin_or_superadmin)
def reissue_equipment(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    
    if not equipment.is_returned:
        messages.error(request, "This equipment is not in returned status.")
        return redirect('equipments:returned')
    
    # Reset return status and related fields
    equipment.is_returned = False
    equipment.return_document = None  # Clear the single document field
    equipment.return_remarks = None
    equipment.return_condition = None
    equipment.return_type = None
    equipment.returned_by = None
    equipment.received_by = None
    
    # Set status back to Active (assuming status ID 1 is Active)
    try:
        active_status = Status.objects.get(name__iexact='Active')
        equipment.status = active_status
    except Status.DoesNotExist:
        # Fallback to first available status
        equipment.status = Status.objects.first()
    
    equipment.updated_by = request.user
    equipment.save()
    
    # Keep return documents in ReturnDocument model for historical purposes
    # Don't delete them - they serve as audit trail
    
    EquipmentActionLog.objects.create(
        equipment=equipment,
        action='reissue',
        user=request.user,
        summary=f"Reissued equipment: {equipment.item_name} (Property #: {equipment.item_propertynum}) - returned to active circulation"
    )
    
    messages.success(request, f"Equipment '{equipment.item_name}' has been reissued and is now available for assignment.")
    return redirect('equipments:returned')

@login_required
@role_required_with_feedback(is_viewer_or_above)
def equipment_history_json(request, equipment_id):
    history = EquipmentHistory.objects.filter(equipment_id=equipment_id).order_by('-changed_at')
    data = [
        {
            'changed_at': h.changed_at.strftime('%Y-%m-%d %H:%M'),
            'action': h.action,
            'field_changed': h.field_changed,
            'old_value': h.old_value,
            'new_value': h.new_value,
            'reason': h.reason if h.reason else 'N/A',  # Show reason or N/A
            'changed_by': h.changed_by.get_full_name() or h.changed_by.username
        }
        for h in history
    ]
    return JsonResponse(data, safe=False)

@login_required
@user_passes_test(is_admin_or_superadmin)
def history_logs(request):
    logs = EquipmentActionLog.objects.select_related('user', 'equipment').order_by('-timestamp')[:500]  # Limit for performance
    return render(request, 'equipments/history_logs.html', {'logs': logs})

@login_required
@user_passes_test(is_admin_or_superadmin)
def clear_history_logs(request):
    if request.method == 'POST':
        # Log the clearing action before actually clearing
        EquipmentActionLog.objects.create(
            equipment=None,
            action='delete',
            user=request.user,
            summary="Cleared all equipment history logs"
        )
        
        EquipmentActionLog.objects.all().delete()
        EquipmentHistory.objects.all().delete()  # Optional: clear field-level history too
        messages.success(request, "All history logs have been cleared.")
    return redirect('equipments:history_logs')

@login_required
@user_passes_test(is_admin_or_superadmin)
def reports_page(request):
    # 1) Pull filter parameters (with sensible defaults)
    date_from = request.GET.get('fromDate')
    date_to   = request.GET.get('toDate')
    period    = request.GET.get('period', 'monthly')  # monthly or yearly

    # 2) Base queryset, then apply date filters if provided
    eqs = Equipment.objects.exclude(item_purdate__isnull=True)
    if date_from:
        eqs = eqs.filter(item_purdate__gte=date_from)
    if date_to:
        eqs = eqs.filter(item_purdate__lte=date_to)

    # 3) Summary values
    from django.db.models import Value as V
    total_asset = eqs.aggregate(total=Sum('item_amount'))['total'] or 0
    categories  = Category.objects.all()
    suppliers   = (
        Equipment.objects
        .exclude(supplier__isnull=True)
        .exclude(supplier='')
        .values_list('supplier', flat=True)
        .distinct()
    )
    selected_category = request.GET.get('category', '')
    selected_supplier = request.GET.get('supplier', '')

    if selected_category:
        eqs = eqs.filter(category_id=selected_category)
    if selected_supplier:
        eqs = eqs.filter(supplier=selected_supplier)

    # 4) Assets by Category & Count
    asset_by_category = (
        eqs.values('category__name')
           .annotate(total=Sum('item_amount'))
           .order_by('-total')
    )
    asset_count_by_category = (
        eqs.values('category__name')
           .annotate(count=Count('id'))
           .order_by('-count')
    )

    # 5) Status breakdown
    status_breakdown = (
        eqs.values('status__name')
           .annotate(count=Count('id'), total=Sum('item_amount'))
           .order_by('-count')
    )

    # 6) Recent assets
    recent_assets = eqs.order_by('-created_at')[:5]

    # 7) Purchases overview (monthly or yearly)
    now = timezone.now()
    current_year = now.year

    monthly_qs = (
        eqs.filter(item_purdate__year=current_year)
           .annotate(period=TruncMonth('item_purdate'))
           .values('period')
           .annotate(count=Count('id'), total=Sum('item_amount'))
           .order_by('period')
    )
    yearly_qs = (
        eqs.annotate(period=ExtractYear('item_purdate'))
           .values('period')
           .annotate(count=Count('id'), total=Sum('item_amount'))
           .order_by('period')
    )

    # 8) Assets by location & status
    assets_by_location_status = (
        eqs.values('location', 'status__name')
           .annotate(count=Count('id'), total=Sum('item_amount'))
           .order_by('location', 'status__name')
    )
    locations = Equipment.objects.values_list('location', flat=True).distinct().order_by('location')
    statuses  = Equipment.objects.values_list('status__name', flat=True).distinct().order_by('status__name')
    sel_loc   = request.GET.get('location', '')
    sel_stat  = request.GET.get('status', '')

    if sel_loc:
        assets_by_location_status = assets_by_location_status.filter(location=sel_loc)
    if sel_stat:
        assets_by_location_status = assets_by_location_status.filter(status__name=sel_stat)

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="finance_report.csv"'
        writer = csv.writer(response)

        # Section: Assets by Category
        writer.writerow(['--- Assets by Category ---'])
        writer.writerow(['Category', 'Total Value (₱)', 'Count'])
        for row in asset_by_category:
            count = next((c['count'] for c in asset_count_by_category if c['category__name'] == row['category__name']), 0)
            writer.writerow([
                row['category__name'],
                f"{row['total']:.2f}" if row['total'] else "0.00",
                count
            ])

        # Section: Purchases Overview
        writer.writerow([])
        writer.writerow(['--- Purchases Overview ---'])
        writer.writerow(['Period', 'Count', 'Total Value (₱)'])
        data = monthly_qs if period == 'monthly' else yearly_qs
        for row in data:
            period_label = row['period'].strftime('%B %Y') if period == 'monthly' else str(row['period'])
            writer.writerow([
                period_label,
                row['count'],
                f"{row['total']:.2f}" if row['total'] else "0.00"
            ])

        # Section: Recently Added Assets
        writer.writerow([])
        writer.writerow(['--- Recently Added Assets ---'])
        writer.writerow(['Property #', 'Name', 'Category', 'Supplier', 'Amount (₱)', 'Purchase Date'])
        for eq in recent_assets:
            writer.writerow([
                eq.item_propertynum,
                eq.item_name,
                eq.category.name if eq.category else '(None)',
                eq.supplier or '(None)',
                f"{eq.item_amount:.2f}" if eq.item_amount else "0.00",
                eq.item_purdate.strftime('%Y-%m-%d') if eq.item_purdate else '(None)'
            ])

        # Section: Assets by Location & Status
        writer.writerow([])
        writer.writerow(['--- Assets by Location & Status ---'])
        writer.writerow(['Location', 'Status', 'Asset Count', 'Total Value (₱)'])
        for row in assets_by_location_status:
            writer.writerow([
                row['location'] or '(None)',
                row['status__name'] or '(None)',
                row['count'],
                f"{row['total']:.2f}" if row['total'] is not None else "0.00"
            ])

        return response


    # 9) CSV export for location/status
    if request.GET.get('export_location_status') == '1':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="assets_by_location_status.csv"'
        writer = csv.writer(response)
        writer.writerow(['Location', 'Status', 'Asset Count', 'Total Value (₱)'])
        for row in assets_by_location_status:
            writer.writerow([
                row['location'] or '(None)',
                row['status__name'] or '(None)',
                row['count'],
                f"{row['total']:.2f}" if row['total'] is not None else "0.00"
            ])
        return response

    # 10) Prepare context
    context = {
        # filter form persistence
        'date_from': date_from,
        'date_to':   date_to,
        'period':    period,

        # dropdowns
        'categories': categories,
        'suppliers':  suppliers,
        'selected_category': selected_category,
        'selected_supplier': selected_supplier,

        # summaries
        'total_asset': total_asset,
        'asset_by_category':       asset_by_category,
        'asset_count_by_category': asset_count_by_category,
        'status_breakdown':        status_breakdown,
        'recent_assets':           recent_assets,

        # purchases overview
        'monthly_purchases': monthly_qs,
        'yearly_purchases':  yearly_qs,

        # location/status grid
        'assets_by_location_status': assets_by_location_status,
        'locations': locations,
        'statuses':  statuses,
        'selected_location': sel_loc,
        'selected_status':  sel_stat,
    }
    return render(request, 'reports/reports.html', context)

@login_required
@user_passes_test(is_admin_or_superadmin)
def generate_report(request):
    from .models import Equipment
    
    # Initialize form only once with GET data
    form = ReportFilterForm(request.GET or None)
    equipments = Equipment.objects.all()
    selected_columns = []  # Ensure selected_columns is always defined

    # --- Advanced Filter Logic ---
    from django.db.models import Q
    from datetime import datetime as dt
    filter_columns = request.GET.getlist('filter_column[]')
    filter_operators = request.GET.getlist('filter_operator[]')
    filter_values = request.GET.getlist('filter_value[]')
    filter_rows = []
    for col, op, val in zip(filter_columns, filter_operators, filter_values):
        filter_rows.append({'column': col, 'operator': op, 'value': val})
    advanced_filters = Q()

    # Field type and lookup map
    field_type_map = {
        'id': 'int',
        'item_amount': 'float',
        'created_at': 'datetime',
        'updated_at': 'datetime',
        'item_purdate': 'date',
        'date_archived': 'datetime',
        'category': 'fk',
        'status': 'fk',
        'is_returned': 'bool',
        'is_archived': 'bool',
    }
    # ForeignKey fields: use exact
    fk_fields = {'category', 'status'}
    # Date/Datetime fields
    date_fields = {'item_purdate', 'created_at', 'updated_at', 'date_archived'}
    # Boolean fields
    bool_fields = {'is_returned', 'is_archived'}
    # Char/Text fields (for icontains)
    char_fields = {
        'item_name', 'item_propertynum', 'item_desc', 'additional_info', 'po_number', 'fund_source', 'supplier',
        'project_name', 'assigned_to', 'end_user', 'location', 'current_location', 'return_remarks', 'return_condition',
        'return_type', 'returned_by', 'received_by',
    }
    # User fields (FK): created_by, updated_by, emp, archived_by
    user_fk_fields = {'created_by', 'updated_by', 'emp', 'archived_by'}

    # Group filters by column to handle OR logic for same column, AND logic between different columns
    column_filters = {}
    
    for col, op, val in zip(filter_columns, filter_operators, filter_values):
        if not col:
            continue
        if op in ['isnull', 'notnull']:
            # Handle null checks separately
            if col not in column_filters:
                column_filters[col] = []
            column_filters[col].append(Q(**{f"{col}__isnull": op == 'isnull'}))
            continue
        if not val and op not in ['isnull', 'notnull']:
            continue
        # Determine lookup and value conversion
        lookup = op
        filter_key = col
        val_conv = val
        # ForeignKey fields: always use exact
        if col in fk_fields:
            lookup = 'exact'
            filter_key = f"{col}__id"
            try:
                val_conv = int(val)
            except Exception:
                continue
        # User FK fields: use exact on id
        elif col in user_fk_fields:
            lookup = 'exact'
            filter_key = f"{col}__id"
            try:
                val_conv = int(val)
            except Exception:
                continue
        # Date/Datetime fields
        elif col in date_fields:
            if op in ['exact', 'gt', 'lt', 'gte', 'lte']:
                filter_key = col + ('__date' if field_type_map.get(col) == 'datetime' else '')
                try:
                    val_conv = dt.strptime(val, '%Y-%m-%d').date()
                except Exception:
                    continue
            else:
                continue  # skip unsupported ops
        # Boolean fields
        elif col in bool_fields:
            lookup = 'exact'
            filter_key = col
            val_conv = val.lower() in ['1', 'true', 'yes', 'on']
        # Char/Text fields
        elif col in char_fields:
            if op not in ['exact', 'icontains']:
                lookup = 'icontains'
            filter_key = col + (f'__{lookup}' if lookup != 'exact' else '')
        else:
            # Fallback: use icontains for unknown fields
            lookup = 'icontains'
            filter_key = col + (f'__{lookup}' if lookup != 'exact' else '')
        # Compose filter and group by column
        if lookup in ['exact', 'iexact', 'icontains', 'gt', 'lt', 'gte', 'lte']:
            if col not in column_filters:
                column_filters[col] = []
            column_filters[col].append(Q(**{filter_key: val_conv}))
    
    # Combine filters: OR within same column, AND between different columns
    for col, filters in column_filters.items():
        if len(filters) == 1:
            advanced_filters &= filters[0]
        else:
            # Multiple filters for same column - use OR
            column_q = Q()
            for f in filters:
                column_q |= f
            advanced_filters &= column_q
    if filter_columns:
        equipments = equipments.filter(advanced_filters)

    if form.is_valid() and not filter_columns:
        if form.cleaned_data['start_date']:
            equipments = equipments.filter(created_at__gte=form.cleaned_data['start_date'])
        if form.cleaned_data['end_date']:
            equipments = equipments.filter(created_at__lte=form.cleaned_data['end_date'])
        if form.cleaned_data['status']:
            equipments = equipments.filter(status=form.cleaned_data['status'])
        if form.cleaned_data['category']:
            equipments = equipments.filter(category=form.cleaned_data['category'])
        if form.cleaned_data['assigned_to']:
            equipments = equipments.filter(assigned_to__icontains=form.cleaned_data['assigned_to'])

    # Calculate total equipment value (needs to be done before exports)
    from django.db.models import Sum
    total_amount = equipments.aggregate(total=Sum('item_amount'))['total'] or 0

    # Handle column selection
    if request.GET.get('columns') or form.is_valid():
        # User has made column selections or form is valid with data
        selected_columns = list(form.cleaned_data.get('columns') or [])
    else:
        # First visit with no parameters - use form's initial values
        selected_columns = form.fields['columns'].initial or []
    
    # If still no columns are selected, use default columns as fallback
    if not selected_columns:
        selected_columns = [
            'user_image',           # Image
            'item_propertynum',     # Property #
            'item_name',            # Name
            'item_desc',            # Description
            'po_number',            # PO Number
            'item_amount',          # Amount
            'end_user',             # End User
            'assigned_to',          # Assigned To
            'category',             # Category
            'item_purdate',         # PO Date
            'current_location',     # Current Location
        ]
    
    seen = set()
    selected_columns = [x for x in selected_columns if not (x in seen or seen.add(x))]

    # Build column_labels from model verbose_name
    from .models import Equipment
    column_labels = {field.name: field.verbose_name.title() for field in Equipment._meta.get_fields() if hasattr(field, 'verbose_name')}
    # Add any missing fields from form choices (for custom/related fields)
    for group, choices in form.fields['columns'].choices:
        for value, label in choices:
            if value not in column_labels:
                column_labels[value] = label

    if request.GET.get('export') == 'csv':
        # For CSV export, use all equipments (before pagination)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="equipment_report.csv"'
        writer = csv.writer(response)
        # Write header
        writer.writerow([column_labels.get(col, col) for col in selected_columns])
        for eq in equipments:  # This is the full queryset before pagination
            row = []
            for col in selected_columns:
                if col == 'category':
                    row.append(str(eq.category) if eq.category else '')
                elif col == 'status':
                    row.append(str(eq.status) if eq.status else '')
                elif col == 'created_at':
                    row.append(eq.created_at.strftime('%Y-%m-%d') if eq.created_at else '')
                else:
                    row.append(getattr(eq, col, ''))
            writer.writerow(row)
        return response
    
    # Debug: Log all export requests
    export_type = request.GET.get('export')
    print(f"DEBUG: Export request received - type: {export_type}")
    
    # Handle Word document export
    if request.GET.get('export') == 'word':
        print("DEBUG: Entering Word export section")
        try:
            from .report_utils import generate_word_report_from_template, create_word_response
            
            # Get user preferences for Word document
            orientation = request.GET.get('orientation', 'portrait')
            table_style = request.GET.get('table_style', 'table_grid')
            font_size = int(request.GET.get('font_size', 10))
            is_preview = request.GET.get('preview') == '1'
            
            # Debug: Print that we're generating Word document
            print(f"DEBUG: Generating Word document with {len(equipments)} equipments")
            print(f"DEBUG: Orientation: {orientation}, Table style: {table_style}, Font size: {font_size}")
            
            # Generate Word document using template with custom options
            doc = generate_word_report_from_template(
                equipments=list(equipments),  # Convert to list for template
                selected_columns=selected_columns,
                column_labels=column_labels,
                filename="equipment_report.docx",
                orientation=orientation,
                table_style=table_style,
                font_size=font_size,
                total_amount=total_amount
            )
            
            # Debug: Confirm document was created
            print(f"DEBUG: Word document created successfully with {len(doc.paragraphs)} paragraphs")
            
            # Create filename with options
            orientation_suffix = "_landscape" if orientation == "landscape" else "_portrait"
            filename = f"WESMAARRDEC_Equipment_Report{orientation_suffix}.docx"
            
            print(f"DEBUG: Returning Word response with filename: {filename}")
            return create_word_response(doc, filename)
            
        except ImportError as e:
            # python-docx not available
            print(f"DEBUG: Import error in Word generation: {e}")
            messages.error(request, f"Word document generation not available: {str(e)}. Downloading CSV instead.")
        except Exception as e:
            # Other errors during Word generation
            print(f"DEBUG: Exception in Word generation: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Word document generation failed: {str(e)}. Downloading CSV instead.")
        
        # Fallback to CSV if Word generation fails
        print("DEBUG: Falling back to CSV due to Word generation error")
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="equipment_report.csv"'
        writer = csv.writer(response)
        writer.writerow([column_labels.get(col, col) for col in selected_columns])
        for eq in equipments:
            row = []
            for col in selected_columns:
                if col == 'category':
                    row.append(str(eq.category) if eq.category else '')
                elif col == 'status':
                    row.append(str(eq.status) if eq.status else '')
                elif col == 'created_at':
                    row.append(eq.created_at.strftime('%Y-%m-%d') if eq.created_at else '')
                elif col == 'updated_at':
                    row.append(eq.updated_at.strftime('%Y-%m-%d') if eq.updated_at else '')
                elif col == 'item_amount':
                    row.append(f"{eq.item_amount:.2f}" if eq.item_amount else '')
                else:
                    value = getattr(eq, col, '')
                    row.append(str(value) if value is not None else '')
            writer.writerow(row)
        return response
    
    # Handle PDF document export
    if request.GET.get('export') == 'pdf':
        print("DEBUG: Entering PDF export section")
        try:
            from .report_utils import generate_pdf_report_from_template, create_pdf_response
            
            # Get user preferences for PDF document
            orientation = request.GET.get('orientation', 'portrait')
            table_style = request.GET.get('table_style', 'table_grid')
            font_size = int(request.GET.get('font_size', 10))
            is_preview = request.GET.get('preview') == '1'
            
            print(f"DEBUG: Generating PDF document with {len(equipments)} equipments")
            
            # Generate PDF document using ReportLab with custom options
            pdf_content = generate_pdf_report_from_template(
                equipments=list(equipments),  # Convert to list for template
                selected_columns=selected_columns,
                column_labels=column_labels,
                filename="equipment_report.pdf",
                orientation=orientation,
                table_style=table_style,
                font_size=font_size
            )
            
            print(f"DEBUG: PDF document created successfully - size: {len(pdf_content)} bytes")
            
            # Create filename with options
            orientation_suffix = "_landscape" if orientation == "landscape" else "_portrait"
            filename = f"WESMAARRDEC_Equipment_Report{orientation_suffix}.pdf"
            
            return create_pdf_response(pdf_content, filename)
        except ImportError as e:
            # ReportLab not available
            print(f"DEBUG: PDF import error: {e}")
            messages.error(request, f"PDF document generation not available: {str(e)}. Downloading CSV instead.")
        except Exception as e:
            # Other errors during PDF generation
            print(f"DEBUG: PDF generation exception: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"PDF document generation failed: {str(e)}. Downloading CSV instead.")
        
        # Fallback to CSV if PDF generation fails
        print("DEBUG: PDF falling back to CSV")
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="equipment_report.csv"'
        writer = csv.writer(response)
        writer.writerow([column_labels.get(col, col) for col in selected_columns])
        for eq in equipments:
            row = []
            for col in selected_columns:
                if col == 'category':
                    row.append(str(eq.category) if eq.category else '')
                elif col == 'status':
                    row.append(str(eq.status) if eq.status else '')
                elif col == 'created_at':
                    row.append(eq.created_at.strftime('%Y-%m-%d') if eq.created_at else '')
                elif col == 'updated_at':
                    row.append(eq.updated_at.strftime('%Y-%m-%d') if eq.updated_at else '')
                elif col == 'item_amount':
                    row.append(f"{eq.item_amount:.2f}" if eq.item_amount else '')
                else:
                    value = getattr(eq, col, '')
                    row.append(str(value) if value is not None else '')
            writer.writerow(row)
        return response
    
    # Add pagination (but not for print_all or export)
    from django.core.paginator import Paginator
    
    # Check if this is a print request for all results
    print_all = request.GET.get('print_all')
    
    if not print_all and request.GET.get('export') != 'csv':
        per_page = request.GET.get('per_page', 25)  # Default to 25 items per page
        try:
            per_page = int(per_page)
            if per_page not in [10, 25, 50, 100]:  # Only allow specific values
                per_page = 25
        except ValueError:
            per_page = 25
        
        paginator = Paginator(equipments, per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    else:
        # For print_all or export, don't paginate - use all results
        from django.core.paginator import Paginator
        paginator = Paginator(equipments, equipments.count() or 1)  # Create a single page with all results
        page_obj = paginator.get_page(1)
    
    # Get all categories for filter dropdown
    categories = Category.objects.all().order_by('name')
    statuses = Status.objects.all().order_by('name')
    end_users = Equipment.objects.exclude(end_user__isnull=True).exclude(end_user='').values_list('end_user', flat=True).distinct().order_by('end_user')
    assigned_to_list = Equipment.objects.exclude(assigned_to__isnull=True).exclude(assigned_to='').values_list('assigned_to', flat=True).distinct().order_by('assigned_to')
    
    # Additional dropdown options
    project_names = Equipment.objects.exclude(project_name__isnull=True).exclude(project_name='').values_list('project_name', flat=True).distinct().order_by('project_name')
    locations = Equipment.objects.exclude(location__isnull=True).exclude(location='').values_list('location', flat=True).distinct().order_by('location')
    current_locations = Equipment.objects.exclude(current_location__isnull=True).exclude(current_location='').values_list('current_location', flat=True).distinct().order_by('current_location')
    fund_sources = Equipment.objects.exclude(fund_source__isnull=True).exclude(fund_source='').values_list('fund_source', flat=True).distinct().order_by('fund_source')
    suppliers = Equipment.objects.exclude(supplier__isnull=True).exclude(supplier='').values_list('supplier', flat=True).distinct().order_by('supplier')
    return_conditions = Equipment.objects.exclude(return_condition__isnull=True).exclude(return_condition='').values_list('return_condition', flat=True).distinct().order_by('return_condition')
    return_types = Equipment.objects.exclude(return_type__isnull=True).exclude(return_type='').values_list('return_type', flat=True).distinct().order_by('return_type')
    returned_by_list = Equipment.objects.exclude(returned_by__isnull=True).exclude(returned_by='').values_list('returned_by', flat=True).distinct().order_by('returned_by')
    received_by_list = Equipment.objects.exclude(received_by__isnull=True).exclude(received_by='').values_list('received_by', flat=True).distinct().order_by('received_by')
    
    # User-related dropdowns
    employees = User.objects.filter(equipment__isnull=False).distinct().order_by('username')
    created_by_users = User.objects.filter(equipment_created__isnull=False).distinct().order_by('username')
    updated_by_users = User.objects.filter(equipment_updated__isnull=False).distinct().order_by('username')
    archived_by_users = User.objects.filter(archived_equipments__isnull=False).distinct().order_by('username')
    
    context = {
        'form': form,
        'equipments': page_obj,  # Changed from equipments to page_obj
        'page_obj': page_obj,    # Add page object for pagination controls
        'selected_columns': selected_columns,
        'column_labels': column_labels,
        'filter_rows': filter_rows or None,  # ensure persistence of filter rows
        'categories': categories,  # pass to template
        'statuses': statuses,
        'end_users': end_users,
        'assigned_to_list': assigned_to_list,
        # Additional dropdown data
        'project_names': project_names,
        'locations': locations,
        'current_locations': current_locations,
        'fund_sources': fund_sources,
        'suppliers': suppliers,
        'return_conditions': return_conditions,
        'return_types': return_types,
        'returned_by_list': returned_by_list,
        'received_by_list': received_by_list,
        'employees': employees,
        'created_by_users': created_by_users,
        'updated_by_users': updated_by_users,
        'archived_by_users': archived_by_users,
        'total_amount': total_amount,
    }
    return render(request, 'reports/generate_report.html', context)


# Document Management Views for Lost Equipment Replacement Documents

@login_required
@secretariat_required
@require_POST
def replace_replacement_document(request, doc_id):
    """Replace an existing replacement document"""
    try:
        # Check if user has permission (admin or encoder)
        if not (is_admin(request.user) or is_encoder(request.user)):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only admins and encoders can replace documents.'
            }, status=403)
            
        # Get the document to replace
        document = get_object_or_404(ReplacementDocument, id=doc_id)
        equipment = document.equipment
        
        # Get the replacement file
        replacement_file = request.FILES.get('replacement_file')
        if not replacement_file:
            return JsonResponse({
                'success': False,
                'error': 'No replacement file provided'
            }, status=400)
            
        # Validate file type and size
        allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png']
        file_extension = '.' + replacement_file.name.split('.')[-1].lower()
        
        if file_extension not in allowed_extensions:
            return JsonResponse({
                'success': False,
                'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
            }, status=400)
            
        if replacement_file.size > 10 * 1024 * 1024:  # 10MB limit
            return JsonResponse({
                'success': False,
                'error': 'File size must be less than 10MB'
            }, status=400)
            
        # Store old filename for logging
        old_filename = document.document.name.split('/')[-1]
        
        # Update the document
        document.document = replacement_file
        document.uploaded_by = request.user
        document.uploaded_at = timezone.now()
        document.save()
        
        # Log the action
        EquipmentActionLog.objects.create(
            equipment=equipment,
            action="Document Replaced",
            details=f"Replaced replacement document '{old_filename}' with '{replacement_file.name}' for Lost equipment",
            user=request.user,
            timestamp=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Document replaced successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@secretariat_required
@require_POST
def delete_replacement_document(request, doc_id):
    """Delete a replacement document"""
    try:
        # Check if user has permission (admin or encoder)
        if not (is_admin(request.user) or is_encoder(request.user)):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only admins and encoders can delete documents.'
            }, status=403)
            
        # Get the document to delete
        document = get_object_or_404(ReplacementDocument, id=doc_id)
        equipment = document.equipment
        filename = document.document.name.split('/')[-1]
        
        # Delete the document
        document.delete()
        
        # Log the action
        EquipmentActionLog.objects.create(
            equipment=equipment,
            action="Document Deleted",
            details=f"Deleted replacement document '{filename}' for Lost equipment",
            user=request.user,
            timestamp=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Document deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@secretariat_required
@require_POST
def add_replacement_documents(request):
    """Add new replacement documents to lost equipment"""
    try:
        # Check if user has permission (admin or encoder)
        if not (is_admin(request.user) or is_encoder(request.user)):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied. Only admins and encoders can add documents.'
            }, status=403)
            
        # Get equipment ID
        equipment_id = request.POST.get('equipment_id')
        if not equipment_id:
            return JsonResponse({
                'success': False,
                'error': 'Equipment ID is required'
            }, status=400)
            
        equipment = get_object_or_404(Equipment, id=equipment_id)
        
        # Check if equipment status is Lost
        if equipment.status.name != 'Lost':
            return JsonResponse({
                'success': False,
                'error': 'Documents can only be added to Lost equipment'
            }, status=400)
            
        # Get uploaded files
        replacement_files = request.FILES.getlist('replacement_files')
        if not replacement_files:
            return JsonResponse({
                'success': False,
                'error': 'No files provided'
            }, status=400)
            
        # Validate and process files
        allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png']
        uploaded_count = 0
        uploaded_files = []
        
        for file in replacement_files:
            # Validate file type and size
            file_extension = '.' + file.name.split('.')[-1].lower()
            
            if file_extension not in allowed_extensions:
                continue  # Skip invalid files
                
            if file.size > 10 * 1024 * 1024:  # 10MB limit
                continue  # Skip large files
                
            # Create replacement document
            doc = ReplacementDocument.objects.create(
                equipment=equipment,
                document=file,
                uploaded_by=request.user,
                uploaded_at=timezone.now()
            )
            uploaded_count += 1
            uploaded_files.append(file.name)
            
        if uploaded_count == 0:
            return JsonResponse({
                'success': False,
                'error': 'No valid files were uploaded. Check file types and sizes.'
            }, status=400)
            
        # Log the action
        file_list = ', '.join(uploaded_files)
        EquipmentActionLog.objects.create(
            equipment=equipment,
            action="Documents Added",
            details=f"Added {uploaded_count} replacement document(s) for Lost equipment: {file_list}",
            user=request.user,
            timestamp=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{uploaded_count} document(s) uploaded successfully',
            'uploaded_count': uploaded_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
